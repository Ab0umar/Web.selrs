from __future__ import annotations

from pathlib import Path
from typing import Dict, List

import openpyxl

INPUT_DIR = Path(r"E:\SELRS.cc\MySQL\Doc\new\normalized")
OUTPUT_DIR = Path(r"E:\SELRS.cc\MySQL\Doc\new\import_ready")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

TARGET_HEADERS: List[str] = [
    "رقم المريض",
    "اسم المريض",
    "تليفون منزل",
    "السن",
    "تاريخ الميلاد",
    "النوع",
    "العنوان",
    "رقم الهوية",
    "كود الطبيب",
    "تاريخ الزيارة",
    "تاريخ فتح الملف",
    "تاريخ الملف",
    "رقم الدكتور",  # keep for auditing
]


def map_row(src: Dict[str, object]) -> Dict[str, object]:
    return {
        "رقم المريض": src.get("رقم المريض"),
        "اسم المريض": src.get("اسم المريض"),
        "تليفون منزل": src.get("تليفون منزل") or src.get("محمول"),
        "السن": src.get("العمر"),
        "تاريخ الميلاد": src.get("تاريخ الميلاد"),
        "النوع": src.get("الجنس"),
        "العنوان": src.get("العنوان"),
        "رقم الهوية": src.get("رقم الهوية"),
        "كود الطبيب": src.get("رقم الدكتور"),
        "تاريخ الزيارة": src.get("تاريخ الملف") or src.get("تاريخ فتح الملف"),
        "تاريخ فتح الملف": src.get("تاريخ فتح الملف"),
        "تاريخ الملف": src.get("تاريخ الملف"),
        "رقم الدكتور": src.get("رقم الدكتور"),
    }


def is_row_empty(values: List[object]) -> bool:
    for v in values:
        if v is None:
            continue
        if str(v).strip() != "":
            return False
    return True


def read_sheet_as_dicts(path: Path) -> List[Dict[str, object]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    header = ["" if c.value is None else str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    while header and header[-1] == "":
        header.pop()

    rows: List[Dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row[: len(header)])
        if is_row_empty(vals):
            continue
        rows.append({header[i]: vals[i] if i < len(vals) else None for i in range(len(header))})

    wb.close()
    return rows


def write_ready(path: Path, rows: List[Dict[str, object]]) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for i, h in enumerate(TARGET_HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)

    out_row = 2
    for src in rows:
        mapped = map_row(src)
        if not str(mapped.get("اسم المريض") or "").strip():
            continue
        for i, h in enumerate(TARGET_HEADERS, start=1):
            ws.cell(row=out_row, column=i, value=mapped.get(h))
        out_row += 1

    wb.save(path)
    wb.close()
    return out_row - 2


def main() -> None:
    report_lines: List[str] = ["Import-Ready Build Report", "=" * 80]
    all_rows: List[Dict[str, object]] = []

    for file in sorted(INPUT_DIR.glob("*_normalized.xlsx")):
        rows = read_sheet_as_dicts(file)
        out_file = OUTPUT_DIR / file.name.replace("_normalized.xlsx", "_import_ready.xlsx")
        count = write_ready(out_file, rows)
        report_lines.append(f"{file.name}: rows={count} -> {out_file.name}")
        all_rows.extend(rows)

    # merged file for single import
    merged_file = OUTPUT_DIR / "all_doctors_import_ready.xlsx"
    merged_count = write_ready(merged_file, all_rows)
    report_lines.append("-" * 80)
    report_lines.append(f"Merged: rows={merged_count} -> {merged_file.name}")

    report_path = OUTPUT_DIR / "import_ready_report.txt"
    report_path.write_text("\n".join(report_lines), encoding="utf-8")

    print("\n".join(report_lines))
    print("=" * 80)
    print(f"Report: {report_path}")


if __name__ == "__main__":
    main()
