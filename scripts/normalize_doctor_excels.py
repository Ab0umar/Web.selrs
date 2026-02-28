from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

TARGET_COLUMNS: List[str] = [
    "رقم المريض",
    "اسم المريض",
    "تاريخ فتح الملف",
    "تاريخ الميلاد",
    "العمر",
    "الجنسية",
    "الجنس",
    "الديانة",
    "رقم الهوية",
    "فصيلة الدم",
    "حالة التدخين",
    "الحالة الاجتماعية",
    "عدد الأولاد",
    "المحافظة",
    "المدينة",
    "العنوان",
    "المهنة",
    "العمل",
    "تليفون منزل",
    "تليفون العمل",
    "محمول",
    "ملاحظات",
    "رقم الملف",
    "تاريخ الملف",
    "ملاحظات.1",
    "رقم الدكتور",
]

ALIASES: Dict[str, str] = {
    "ملاحظات1": "ملاحظات.1",
    "ملاحظات 1": "ملاحظات.1",
    "رقم دكتور": "رقم الدكتور",
    "رقم الطبيب": "رقم الدكتور",
}

# Fallback mapping from filename stems to doctor number
FILENAME_DOCTOR_MAP: Dict[str, int] = {
    "الصواف": 5,
    "عيسي": 4,
    "احمد رشدي": 8,
    "السعدني": 1,
    "حازم": 6,
    "حازم ": 6,
    "سعيد": 12,
    "السيد": 7,
    "كريم": 3,
    "محمد عبدالرحيم": 2,
}

SKIP_FILES = {
    "missing_patient_codes_vs_all.xlsx",  # no reliable header
}


def normalize_col_name(name: str) -> str:
    k = (name or "").strip()
    return ALIASES.get(k, k)


def detect_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, max_rows: int = 20) -> Optional[Tuple[int, Dict[str, int]]]:
    for rix, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        values = ["" if v is None else str(v).strip() for v in row]
        norm = [normalize_col_name(v) for v in values]
        present = {v: i for i, v in enumerate(norm) if v}

        match_count = sum(1 for col in TARGET_COLUMNS if col in present)
        if match_count >= 4 and ("رقم المريض" in present or "اسم المريض" in present):
            return rix, present
    return None


def get_first_sheet(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    return wb, ws


def is_empty_row(values: List[object]) -> bool:
    for v in values:
        if v is None:
            continue
        if str(v).strip() != "":
            return False
    return True


def coerce_int(value: object) -> Optional[int]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def process_file(path: Path, out_dir: Path) -> Tuple[str, int, str]:
    if path.name in SKIP_FILES:
        return path.name, 0, "SKIPPED (configured)"

    wb_in, ws_in = get_first_sheet(path)
    detected = detect_header_row(ws_in)
    if not detected:
        wb_in.close()
        return path.name, 0, "SKIPPED (header not found)"

    header_row, src_idx = detected

    # discover doctor id from existing rows if present
    discovered_ids = set()
    doctor_col = src_idx.get("رقم الدكتور")
    if doctor_col is not None:
        for row in ws_in.iter_rows(min_row=header_row + 1, values_only=True):
            if doctor_col < len(row):
                did = coerce_int(row[doctor_col])
                if did is not None:
                    discovered_ids.add(did)
    inferred_doctor = next(iter(discovered_ids)) if len(discovered_ids) == 1 else None
    fallback_doctor = FILENAME_DOCTOR_MAP.get(path.stem)

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet1"

    for cix, col in enumerate(TARGET_COLUMNS, start=1):
        ws_out.cell(row=1, column=cix, value=col)

    out_row = 2
    for row in ws_in.iter_rows(min_row=header_row + 1, values_only=True):
        if is_empty_row(list(row)):
            continue

        row_out = []
        has_key_data = False
        for col in TARGET_COLUMNS:
            src_col = src_idx.get(col)
            val = row[src_col] if src_col is not None and src_col < len(row) else None
            row_out.append(val)
            if col in ("رقم المريض", "اسم المريض") and val not in (None, ""):
                has_key_data = True

        if not has_key_data:
            continue

        # fill missing doctor number
        did_idx = TARGET_COLUMNS.index("رقم الدكتور")
        did = coerce_int(row_out[did_idx])
        if did is None:
            did = inferred_doctor if inferred_doctor is not None else fallback_doctor
            if did is not None:
                row_out[did_idx] = did

        for cix, val in enumerate(row_out, start=1):
            ws_out.cell(row=out_row, column=cix, value=val)
        out_row += 1

    out_path = out_dir / f"{path.stem}_normalized.xlsx"
    wb_out.save(out_path)

    wb_in.close()
    wb_out.close()

    return path.name, out_row - 2, f"OK -> {out_path.name}"


def main() -> None:
    parser = argparse.ArgumentParser(description="Normalize doctor Excel files to unified schema")
    parser.add_argument("--input", default=r"E:\SELRS.cc\MySQL\Doc\new", help="Input folder containing xlsx files")
    parser.add_argument("--output", default=r"E:\SELRS.cc\MySQL\Doc\new\normalized", help="Output folder")
    args = parser.parse_args()

    in_dir = Path(args.input)
    out_dir = Path(args.output)
    out_dir.mkdir(parents=True, exist_ok=True)

    results: List[Tuple[str, int, str]] = []
    for file_path in sorted(in_dir.glob("*.xlsx")):
        if file_path.name.endswith("_normalized.xlsx"):
            continue
        results.append(process_file(file_path, out_dir))

    report_lines = ["Normalization Report", "=" * 80]
    for name, rows, status in results:
        report_lines.append(f"{name}: rows={rows} | {status}")

    report_path = out_dir / "normalization_report.txt"
    report_path.write_text("\n".join(report_lines), encoding="utf-8")

    print("\n".join(report_lines))
    print("=" * 80)
    print(f"Report: {report_path}")


if __name__ == "__main__":
    main()
